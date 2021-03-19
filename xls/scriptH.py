import openpyxl
import re
import json
wb = openpyxl.load_workbook('./stanPoNaseljima2011Hrvatska.xlsx')

# grab the active worksheet
sheet = wb.active
output = {}
#1col : 3col
for i in range(1,sheet.max_row):
    name = str(sheet.cell(row=i,column =1).value).strip().lower()
    stanovnistvo = sheet.cell(row=i,column=3).value
    if(name != '' and name != ' '):
        nameNoSpace = re.sub(' ','',name)
        output[nameNoSpace] = {}
        output[nameNoSpace]['brojStanovnika'] = stanovnistvo
        output[nameNoSpace]['originalNaziv'] = name

with open('gradoviH.json','w',encoding="UTF-8") as f:
    json.dump(output,f,ensure_ascii=False)

